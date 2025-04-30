from openpyxl import load_workbook

# Abrir o arquivo Excel
wb = load_workbook("ControleGastos.xlsx")
ws = wb.active

# Mapeamento dos meses e colunas
colunas_meses = {
    "ABR": "B", "MAI": "C", "JUN": "D", "JUL": "E", 
    "AGO": "F", "SET": "G", "OUT": "H", "NOV": "I", "DEZ": "J"
}

# Itens com as linhas correspondentes na planilha
itens = {
    "Moradia": 2,
    "Transporte": 3,
    "Internet": 4,
    "Assinaturas": 5,
    "Mercado": 6,  # Este será descontado do vale-refeição
    "Saúde": 7,
    "Lazer": 8,
    "Educação": 9
}

# Mostrar opções de mês
print("\nMeses disponíveis:")
for mes in colunas_meses:
    print(f"- {mes}")

# Escolher o mês
mes_escolhido = input("\nDigite o mês que deseja preencher (ex: JUN): ").strip().upper()
coluna = colunas_meses.get(mes_escolhido)

if not coluna:
    print("Mês inválido. Verifique e tente novamente.")
    exit()

print(f"\n>>> Preenchendo os dados para o mês: {mes_escolhido}\n")

# Totais separados
total_salario = 0
total_vale = 0

for nome_item, linha in itens.items():
    entrada = input(f"{nome_item}: ").strip()

    try:
        valor_numerico = float(entrada.replace("R$", "").replace(",", "."))
        valor = None if valor_numerico == 0 else valor_numerico

        if nome_item == "Mercado":
            total_vale += valor_numerico
        else:
            total_salario += valor_numerico
    except:
        valor = "-"

    if linha != 10:  # linha 10 = TOTAL, mantém a fórmula
        ws[f"{coluna}{linha}"] = valor

# Mostrar totais estimados
print(f"\nTotal estimado do salário (sem mercado): R$ {total_salario:.2f}")
print(f"Total estimado do vale (somente mercado): R$ {total_vale:.2f}\n")

# Saldo salário (linha 14)
salario_input = input("Salário recebido neste mês: R$ ").replace(",", ".")
try:
    salario = float(salario_input)
    saldo = salario - total_salario
    ws[f"{coluna}14"] = saldo
except:
    ws[f"{coluna}14"] = "-"

# Vale-refeição (linha 15)
vale = input("Vale Refeição: R$ ").strip()
try:
    valor_vale = float(vale.replace("R$", "").replace(",", "."))
    saldo_vale = valor_vale - total_vale
    ws[f"{coluna}15"] = saldo_vale
except:
    ws[f"{coluna}15"] = "-"

# Salvar o arquivo
wb.save("ControleGastos.xlsx")
print("\n✅ Planilha preenchida com sucesso! Arquivo salvo como 'ControleGastos.xlsx'")
